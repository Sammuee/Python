# Excel Spreadsheets
import openpyxl as xl
from openpyxl.chart import BarChart, Reference      # importing two classes "BarChart" and "Reference"


def process_workboot(input_file, output_file):
    input_path = f'C:\\Users\\Sammue\\Desktop\\Python\\Input\\'
    output_path = f'C:\\Users\\Sammue\\Desktop\\Python\\Output\\'
    wb = xl.load_workbook(f'{input_path}{input_file}.xlsx')        # will create a workbook object based on the excel file in current path
    sheet = wb['Worksheet']                     # Access the worksheet


    for row in range(2, sheet.max_row + 1):         # max rows in excel worksheet; starting by the second row (header) and will iterate till row 4 (4+1 needed)
        cell = sheet.cell(row, 3)                   # will access the worksheet for each row from (2-4) and column 3
        corrected_price = cell.value * 0.9          # reduce the value of each cell by 10%
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                        min_row=2,
                        max_row=sheet.max_row,      # select the rows from row 2 to 4 (max row)
                        min_col=4,
                        max_col=4)                  # select only column 4 from the above mentioned rows
    # 'values' now contains the name of the Worksheet incl. the referenced cells: 'Worksheet'!$D$2:$D$4
    chart = BarChart()                              # creates a BarChart object (from class)
    chart.add_data(values)                          # calls method and adds data from 'values' (which are our referenced cells)
    sheet.add_chart(chart, 'B6')                    # adds the chart to cell 'E2'

    wb.save(f'{output_path}{output_file}.xlsx')                      # saves excel file
    print(f"Excel file: {output_path}{output_file}.xlsx saved.")


excel_file_name = input("Which excel file do you want to amend? ")
excel_target_name = input("Filename for amended excel? ")

process_workboot(excel_file_name, excel_target_name)
